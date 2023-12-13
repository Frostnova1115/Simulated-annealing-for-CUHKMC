# importing packets
import os
import math
import copy
import random
import openpyxl
import datetime
import type_def
import functools
import param_init
import numpy as np
import pandas as pd
from openpyxl.styles import Font
from matplotlib import pyplot as plt
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

room_num = 59
near_nurse_room_num = 8
bed_num = 144
department_to_floor = {"Medical": ["C7"], "Mixed": ["C8"], "Ortho": ["B9"], "Surgical": ["A10", "B10"]}
floor_to_department = {"C7": "Medical", "C8": "Mixed", "B9": "Ortho", "A1": "Surgical", "B1": "Surgical"}

def dayCount(date, move):
    timestamp = datetime.datetime.strptime(date, "%Y/%m/%d")
    timestamp += datetime.timedelta(days = move)
    return timestamp.strftime("%Y/%m/%d")


def insert(patient, raw_schedule):
    schedule = copy.deepcopy(raw_schedule)
    for i in schedule:
        if i > "A1006 4":
            break
        availibility = 1
        if True:
        # if patient.department == floor_to_department[i[0:2]]:
            # print(patient.department,' ', floor_to_department[i[0:2]])
            for j in schedule[i]:
                if clash(patient, j):
                    availibility = 0
            for k in schedule:
                if k.split()[0] == i.split()[0]:
                    for j in schedule[k]:
                        if clash(patient, j) and patient.gender != j.gender:
                            availibility = 0
            if availibility == 1:
                return i
    return -1


def objectiveFunction(schedule, eval = 0):
    table = {}
    score = 0
    preference = 0
    for i in range(1,31):
        table[dayCount(begin_date, i-1)] = dict([(i, 0) for i in schedule])
    # score for room occupancy:
    for i in schedule:
        for j in schedule[i]:
            date = j.admission
            while date <= j.discharge:
                table[date][i] = 1
                date = dayCount(date, 1)
    for i in table:
        date = len(table)
        room = 'A1002'
        count = 0
        for j in table[i]:
            if room not in j:
                capacity = Rooms[id_to_index[room]].capacity
                if count == 0:
                    score += 2
                else:
                    score += int(2.0*(float(count)/float(capacity)))
                room = j.split()[0]
                count = table[i][j]
            else:
                count += table[i][j]
        date -= 1
    # score for preference
    for i in schedule:
        for j in schedule[i]:
            room = i.split()[0]
            if Rooms[id_to_index[room]].capacity == j.preference:
                score += 20
                preference += 1
    # score for nurse station
    if eval:
        print("preferences: ", preference)
    return -score - 10 * preference

# testing obj
# objectiveFunction(Schedule)
# print(objectiveFunction(Schedule))
# 1. move one patient directly

def clash(a, b):
    if a.admission <= b.admission and b.admission <= a.discharge:
        return 1
    if a.admission <= b.discharge and b.discharge <= a.discharge:
        return 1
    if b.admission <= a.admission and a.admission <= b.discharge:
        return 1
    if b.admission <= a.discharge and a.discharge <= b.discharge:
        return 1
    return 0

def method_1(raw_schedule):
    schedule = copy.deepcopy(raw_schedule)
    patients_num = len(Patients)
    move = random.randint(1, patients_num)
    patient = -1
    pre_bed = -1
    pre_room = -1
    for i in schedule:
        for j in schedule[i]:
            move -= 1
            if move == 0:
                patient = j
                pre_bed = i
                pre_room = i.split()[0]
                schedule[i].remove(j)
    # find availbe beds
    for k in Rooms:
        if k.id == pre_room:
            k.number -= 1
            if k.number == 0:
                k.gender = ""
    empty_beds = []
    for i in schedule:
        # small sample
        if i > "A1006 4":
            break
        availbility = 1
        # capacity constraint
        for j in schedule[i]:
            if clash(j, patient):
                availbility = 0
        # gender constraint
        for j in schedule:
            if j.split()[0] == i.split()[0]:
                for k in schedule[j]:
                    if k.gender != patient.gender:
                        if k.discharge > patient.admission and k.admission < patient.discharge:
                            availbility = 0
                            break
        # department
        # if floor_to_department[i[0:2]] != patient.department and i[0:2] != "C8":
        #     availbility = 0
        if availbility == 1:
            empty_beds.append(i)
    enter = empty_beds[random.randint(1, len(empty_beds))-1]
    schedule[enter].append(patient)
    for k in Rooms:
        if k.id == pre_room:
            k.number += 1
            if k.number == 1:
                k.gender = patient.gender
    return schedule

def check_exchange(raw_schedule, a, b):
    schedule = copy.deepcopy(raw_schedule)
    count = 0
    patient_a = ""
    patient_b = ""
    bed_a = ""
    bed_b = ""
    room_a = ""
    room_b = ""
    # find a and remove
    for i in schedule:
        for j in schedule[i]:
            count += 1
            if count == a:
                bed_a = i
                patient_a = j
                for k in Rooms:
                    if k.id == bed_a.split()[0]:
                        room_a = copy.deepcopy(k)
    # find b and remove
    count = 0
    for i in schedule:
        for j in schedule[i]:
            count += 1
            if count == b:
                bed_b = i
                patient_b = j
                for k in Rooms:
                    if k.id == bed_b.split()[0]:
                        room_b = copy.deepcopy(k)
    schedule[bed_a].remove(patient_a)
    schedule[bed_b].remove(patient_b)
    # same department
    # if patient_a.department != patient_b.department and (room_a.id.find("C8") == -1 or room_b.id.find("C8") == -1):
    #     return 0
    # gender policy
    for j in schedule:
        if j.split()[0] == bed_a.split()[0]:
            for k in schedule[j]:
                if k.gender != patient_b.gender:
                    if clash(k, patient_b):
                        return 0
    for j in schedule:
        if j.split()[0] == bed_b.split()[0]:
            for k in schedule[j]:
                if k.gender != patient_a.gender:
                    if clash(k, patient_a):
                        return 0
    # time clash in bed a
    for i in schedule[bed_a]:
        if clash(i, patient_b):
            return 0
    # time clash in bed b
    for i in schedule[bed_b]:
        if clash(i, patient_a):
            return 0
    return 1

def exchange(raw_schedule, a, b):
    schedule = copy.deepcopy(raw_schedule)
    count = 0
    patient_a = ""
    patient_b = ""
    bed_a = ""
    bed_b = ""
    # find a and remove
    for i in schedule:
        for j in schedule[i]:
            count += 1
            if count == a:
                bed_a = i
                patient_a = j
    # find b and remove
    count = 0
    for i in schedule:
        for j in schedule[i]:
            count += 1
            if count == b:
                bed_b = i
                patient_b = j
    schedule[bed_a].remove(patient_a)
    schedule[bed_b].remove(patient_b)
    schedule[bed_a].append(patient_b)
    schedule[bed_b].append(patient_a)
    return schedule

def method_2(raw_schedule):
    schedule = copy.deepcopy(raw_schedule)
    exchange_list = []
    for i in range(0,len(Patients)):
        for j in range(i+1,len(Patients)):
            if check_exchange(schedule, i+1, j+1):
                exchange_list.append((i+1, j+1))
    if len(exchange_list) > 0:
        pair = exchange_list[random.randint(1,len(exchange_list))-1]
        return exchange(schedule, pair[0], pair[1])
    else:
        return raw_schedule

def neighborFunction(raw_schedule):
    # randomly sclect a way to generate neighbour
    schedule = copy.deepcopy(raw_schedule)
    method = random.randint(1, 2)
    new_schedule = copy.deepcopy(schedule)
    # 1. move one patient
    if method == 1:
        new_schedule = method_1(schedule)
    # swap two patients
    elif method == 2:
        new_schedule = method_2(schedule)
    return new_schedule


def simulatedAnnealing(currentSchedule):
    # temperature
    T = 1000
    # cooling rate
    k = 0.9
    # max step number
    maxStep = 25
    # step finished
    step = 0
    # initial solution
    schedule = copy.deepcopy(currentSchedule)
    # current best solution* date
    solution = copy.deepcopy(currentSchedule)
    while step < maxStep:
        temp_schedule = copy.deepcopy(schedule)
        for i in range(0, 20):
            # find neighbor
            new_schedule = neighborFunction(temp_schedule)
            # find neighbor's energy
            E_new = objectiveFunction(new_schedule)
            # find temp's energy
            E_temp = objectiveFunction(temp_schedule)
            delta =  E_new - E_temp
            if delta < 0:
                # accept good solution
                temp_schedule = copy.deepcopy(new_schedule)
                # update best solution if possible
                # find best solution's energy
                E_solution = objectiveFunction(solution)
                if E_solution > E_temp:
                    solution = copy.deepcopy(temp_schedule)
            else:
                # accept bad solution with probability p
                p = math.exp(-delta/T)
                # print(p)
                pull = random.random()
                if pull < p:
                    temp_schedule = copy.deepcopy(new_schedule)
            # update current schedule
            E_schedule = objectiveFunction(schedule)
            if E_schedule > E_temp:
                schedule = copy.deepcopy(temp_schedule)
        # update parameters
        step += 1
        T = k * T
    return solution

def outPutSchedule(current_schedule, admission_num, exe_time):
    # workbook = openpyxl.load_workbook("testing.xlsx")
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    red = PatternFill(start_color="00FF6D6D", end_color="00FF6D6D", fill_type="solid")
    blue = PatternFill(start_color="008FAAFF", end_color="008FAAFF", fill_type="solid")
    yellow = PatternFill(start_color="00A5FF5B", end_color="00A5FF5B", fill_type="solid")
    date_to_column = {}
    id_to_row = {}

    # print date row
    for i in range(1, 31):
        sheet.cell(row = 1, column = i + 1, value = dayCount(begin_date, i - 1)).alignment = Alignment(horizontal='center', vertical='center')
        date_to_column[dayCount(begin_date, i - 1)] = i + 1

    # print room number column
    # for i in range(1,len(Beds) + 1):
    for i in range(1, 17):
        sheet.cell(row = i + 1, column = 1, value = Beds[i - 1].id).alignment = Alignment(horizontal='center', vertical='center')
        id_to_row[Beds[i-1].id] = i+1

    # set width and height
    for i in range(1, sheet.max_row + 1):
        sheet.row_dimensions[i].height = 15
    for i in range(1,sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(i)].width = 12.5
    # load schedule
    occupied = 0
    for i in current_schedule:
        for j in current_schedule[i]:
            row_num = id_to_row[i]
            start = date_to_column[j.admission]
            end = date_to_column[j.discharge]
            date_stamp = start
            if j.admission < Date_now:
                color = red
            else:
                color = blue
            while date_stamp <= end:
                occupied += 1
                if date_stamp == start:
                    sheet.cell(row = row_num, column = date_stamp, value = j.id).fill = color
                else:
                    sheet.cell(row = row_num, column = date_stamp).fill = color
                date_stamp += 1
    # occupancy
    sheet.cell(row = 18, column = 2, value = str(round(float(occupied)/(32*30)*100, 2)) + '%')
    sheet.cell(row = 18, column = 1, value = "Occupancy").alignment = Alignment(horizontal='center', vertical='center')
    # admission rate
    sheet.cell(row = 19, column = 2, value = str(round(float(admission_num)/(50)*100, 2)) + '%')
    sheet.cell(row = 19, column = 1, value = "Admission").alignment = Alignment(horizontal='center', vertical='center')
    # execution time
    sheet.cell(row = 20, column = 2, value = exe_time)
    sheet.cell(row = 20, column = 1, value = "Time").alignment = Alignment(horizontal='center', vertical='center')

    print("Occupancy = ", str(round(float(occupied)/(32*30)*100, 2)) + '%')
    workbook.save(filename="Schedules_floating\Schedule" + str(dataset) + ".xlsx")
    workbook.close()

def optimize(raw_schedule, round_num):
    schedule = copy.deepcopy(raw_schedule)
    st = datetime.datetime.now()
    count = 0
    while count < 55 and len(WaitingList) > 0:
        # test SA
        count += 1
        # pos = random.randint(0,len(WaitingList)-1)
        pos = 0
        bed = insert(WaitingList[pos],schedule)
        if bed != -1:
            Patients.append(WaitingList[pos])
            schedule[bed].append(WaitingList[pos])
            del WaitingList[pos]
        # print(schedule)
        schedule = simulatedAnnealing(schedule)
        # for i in schedule:
        #     if len(schedule[i]) > 1:
        #         print("Warning")
        ed = datetime.datetime.now()
        exe_time = (ed-st).seconds
        if count <= 1:
            print("Dataset number: ", round_num)
        if len(WaitingList) == 0 or count == 55:
            print("number: ", count)
            print("bed = ", bed)
            print("Length of WL: ", len(WaitingList))
            print("evaluation: ", objectiveFunction(schedule, eval = 1))
            print("execution time = ", exe_time)
            outPutSchedule(schedule, len(Patients), exe_time)
            print("----------------")
        # if exe_time > 5:
        #     print("Saving current schedule...")
        #     with open("inline_data\schedule.txt", "w") as f:
        #         f.write(str(schedule))
        #     print("Save completed.")
        #     print("Saving waiting list...")
        #     with open("inline_data\waitinglist.txt") as f:
        #         f.write(WaitingList)
        #     print("Waiting list saved")
        #     print("TLE")
        #     break
    return schedule

Departments = []
Rooms = []
Beds = []
Patients = []
WaitingList = []
begin_date = ''
dataset = 0
for dataset in range(1, 11):
    # initialization parameters
    # param_init.init_Patients(dataset)
    Departments = copy.deepcopy(param_init.Departments)
    Rooms = copy.deepcopy(param_init.Rooms)
    Beds = copy.deepcopy(param_init.Beds)
    Patients = []
    WaitingList = copy.deepcopy(param_init.init_Patients(dataset))
    begin_date = dayCount(min([i.admission for i in WaitingList]), 0)
    Date_now = begin_date
    # Room id to idex
    id_to_index = {}
    for k in range(0, len(Rooms)):
        id_to_index[Rooms[k].id] = k
    Schedule = dict([(Beds[i].id, []) for i in range(len(Beds))])
    Schedule = optimize(Schedule, dataset)
# Schedule = optimize(Schedule)
